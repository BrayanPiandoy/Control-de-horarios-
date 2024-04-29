import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

try:
    # Obtener la ruta del escritorio del usuario
    escritorio = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

    # Ruta completa del archivo de la base de datos en el escritorio
    base_datos_ruta = os.path.join(escritorio, 'app_horarios', 'horarios.db')

    # Ruta completa del archivo del reporte en el escritorio
    reporte_ruta = os.path.join(escritorio, 'reporte_horarios.xlsx')

    # Verificar si el archivo de reporte ya existe
    if os.path.exists(reporte_ruta):
        # Si el archivo existe, cargar el libro existente
        wb = load_workbook(filename=reporte_ruta)
        ws = wb.active
        
        # Verificar si hay datos en el archivo existente
        if ws.max_row > 1:
            # Crear una lista de identificadores únicos
            ids_existentes = set([row[0].value for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row)])
        else:
            # Si no hay datos en el archivo existente, inicializar la lista de IDs existentes como vacía
            ids_existentes = set()
    else:
        # Si el archivo no existe, crear un nuevo libro de Excel y una hoja de trabajo
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Horarios"
        # Agregar encabezados de columna
        ws.append(["Num_Registro", "Nombre", "Hora de Entrada", "Hora de Salida", "Computador", "Fecha"])
        # Inicializar la lista de IDs existentes como vacía
        ids_existentes = set()

    # Conectar a la base de datos SQLite
    conn = sqlite3.connect(base_datos_ruta)
    c = conn.cursor()

    # Verificar si la tabla horarios existe
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='horarios'")
    tabla_existe = c.fetchone()

    if tabla_existe:
        # Ejecutar una consulta para obtener los datos que deseas guardar en Excel
        c.execute("SELECT * FROM horarios")
        data = c.fetchall()

        # Agregar datos de la base de datos a la hoja de trabajo solo si el ID no está presente
        for row in data:
            if row[0] not in ids_existentes:
                ws.append(row)
                ids_existentes.add(row[0])  # Agregar el nuevo ID a la lista de IDs existentes

                # Aplicar formato de centrado a todas las celdas en la fila recién agregada
                for cell in ws.iter_rows(max_row=ws.max_row):
                    for c in cell:
                        c.alignment = Alignment(horizontal='center', vertical='center')

        # Ajustar el ancho de las columnas y el alto de las filas para que sean más grandes
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Ajustar el ancho de la columna
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        for row in ws.rows:
            ws.row_dimensions[row[0].row].height = 20  # Ajustar la altura de la fila

        # Aplicar bordes a todas las celdas de la tabla
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                thin_border = Border(left=Side(style='thin'), 
                                     right=Side(style='thin'), 
                                     top=Side(style='thin'), 
                                     bottom=Side(style='thin'))
                cell.border = thin_border

        # Aplicar un color claro al encabezado
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        for cell in ws[1]:
            cell.fill = header_fill

        # Guardar el libro de Excel en el escritorio
        wb.save(filename=reporte_ruta)

        print(f"El reporte se ha actualizado correctamente en {reporte_ruta}")
    else:
        print("La tabla 'horarios' no existe en la base de datos.")

except sqlite3.Error as error:
    print("Error al trabajar con la base de datos:", error)

finally:
    # Cerrar la conexión a la base de datos
    if conn:
        conn.close()
